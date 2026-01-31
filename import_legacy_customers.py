#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
IMPORTACIÓN DE CLIENTES LEGACY A ODOO
================================================================================

Este script importa clientes desde un Excel del sistema legacy hacia Odoo.

CAMPOS MAPEADOS:
- Número       → ref (referencia interna)
- Nombre       → name
- Domicilio    → street
- Localidad    → city
- CP           → zip
- Teléfonos    → phone
- CUIT         → vat (formateado para Argentina)
- IVA          → l10n_ar_afip_responsibility_type_id
- Mail         → email

CARACTERÍSTICAS:
- Detecta y omite clientes que ya existen (por CUIT o ref)
- Mapea tipos de IVA argentinos
- Es idempotente
- Soporta dry-run

USO:
    python import_legacy_customers.py --excel /path/to/clientes.xlsx --dry-run
    python import_legacy_customers.py --excel /path/to/clientes.xlsx --execute

AUTOR: Generado para Mundo Limpio - Migración Odoo 18
FECHA: 2026-01-31
================================================================================
"""

import argparse
import logging
import os
import re
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
import xmlrpc.client

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

ODOO_URL = os.getenv("ODOO_URL", "https://mundolimpio.aftermoves.com")
ODOO_DB = os.getenv("ODOO_DB", "Testing")
ODOO_USER = os.getenv("ODOO_USER", "fabriziodominguez@aftermoves.com")
ODOO_PASSWORD = os.getenv("ODOO_PASSWORD", "admin")

# Número de hilos para importación paralela
NUM_THREADS = int(os.getenv("NUM_THREADS", "10"))

# Mapeo de tipos de IVA del sistema legacy a Odoo Argentina
# l10n_ar_afip_responsibility_type_id
IVA_TYPE_MAP = {
    'RI': 'IVA Responsable Inscripto',
    'RNI': 'IVA Responsable No Inscripto', 
    'M': 'Responsable Monotributo',
    'MT': 'Responsable Monotributo',
    'MONO': 'Responsable Monotributo',
    'EX': 'IVA Sujeto Exento',
    'EXENTO': 'IVA Sujeto Exento',
    'CF': 'Consumidor Final',
    'CONS': 'Consumidor Final',
    'NC': 'IVA No Alcanzado',
    'NA': 'IVA No Alcanzado',
}

# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class LegacyCustomer:
    """Representa un cliente del sistema legacy"""
    row_number: int
    code: str = ""
    name: str = ""
    street: str = ""
    city: str = ""
    zip_code: str = ""
    phone: str = ""
    cuit: str = ""
    iva_type: str = ""
    email: str = ""
    
    # Procesamiento
    is_valid: bool = True
    validation_errors: List[str] = field(default_factory=list)
    
    @property
    def vat_clean(self) -> str:
        """Retorna el CUIT/DNI sin guiones, solo números"""
        if not self.cuit:
            return ""
        # Limpiar el CUIT - solo números
        cuit_clean = re.sub(r'[^0-9]', '', self.cuit)
        if len(cuit_clean) >= 7:  # DNI mínimo 7 dígitos, CUIT 11
            return cuit_clean
        return ""
    
    @property
    def is_cuit(self) -> bool:
        """Determina si es CUIT (11 dígitos) o DNI (7-8 dígitos)"""
        return len(self.vat_clean) == 11
    
    @property
    def is_dni(self) -> bool:
        """Determina si es DNI"""
        return 7 <= len(self.vat_clean) <= 8


@dataclass
class ImportResult:
    """Resultado de la importación"""
    dry_run: bool = True
    total_rows: int = 0
    valid_customers: int = 0
    invalid_customers: int = 0
    customers_created: int = 0
    customers_updated: int = 0
    customers_skipped: int = 0
    errors: List[str] = field(default_factory=list)
    created_ids: List[int] = field(default_factory=list)
    log_entries: List[str] = field(default_factory=list)


# =============================================================================
# PARSER DE EXCEL
# =============================================================================

class LegacyCustomerParser:
    """Parser para Excel de clientes legacy"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.customers: List[LegacyCustomer] = []
        
    def parse(self) -> List[LegacyCustomer]:
        """Parsea el Excel y retorna lista de clientes"""
        logger.info(f"Parseando archivo: {self.file_path}")
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Instale openpyxl: pip install openpyxl")
        
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        # Buscar fila de encabezados (contiene "Número", "Nombre", etc.)
        header_row = None
        col_map = {}
        
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            row_text = " ".join(str(c).lower() if c else "" for c in row)
            if 'número' in row_text or 'numero' in row_text:
                header_row = idx
                # Mapear columnas
                for col_idx, cell in enumerate(row):
                    if cell:
                        cell_lower = str(cell).lower().strip()
                        if 'núm' in cell_lower or 'num' in cell_lower:
                            col_map['code'] = col_idx
                        elif 'nombre' in cell_lower:
                            col_map['name'] = col_idx
                        elif 'domic' in cell_lower or 'direc' in cell_lower:
                            col_map['street'] = col_idx
                        elif 'local' in cell_lower or 'ciudad' in cell_lower:
                            col_map['city'] = col_idx
                        elif 'cp' in cell_lower or 'postal' in cell_lower:
                            col_map['zip'] = col_idx
                        elif 'tel' in cell_lower:
                            col_map['phone'] = col_idx
                        elif 'cuit' in cell_lower:
                            col_map['cuit'] = col_idx
                        elif 'iva' in cell_lower:
                            col_map['iva'] = col_idx
                        elif 'mail' in cell_lower or 'email' in cell_lower:
                            col_map['email'] = col_idx
                break
        
        if not header_row:
            raise ValueError("No se encontró fila de encabezados en el Excel")
        
        logger.info(f"Encabezados encontrados en fila {header_row}: {col_map}")
        
        # Parsear datos
        for idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            customer = self._parse_row(row, idx, col_map)
            if customer:
                self.customers.append(customer)
        
        wb.close()
        
        valid = len([c for c in self.customers if c.is_valid])
        invalid = len([c for c in self.customers if not c.is_valid])
        logger.info(f"Parseados {len(self.customers)} clientes: {valid} válidos, {invalid} inválidos")
        
        return self.customers
    
    def _parse_row(self, row: tuple, row_number: int, col_map: Dict[str, int]) -> Optional[LegacyCustomer]:
        """Parsea una fila y retorna un cliente"""
        # Verificar si es fila de datos (tiene código numérico)
        code_idx = col_map.get('code', 0)
        if code_idx >= len(row) or not row[code_idx]:
            return None
        
        code_val = str(row[code_idx]).strip()
        # Limpiar .0 de floats
        code_val = code_val.replace('.0', '')
        
        if not code_val or not code_val.replace('.', '').isdigit():
            return None
        
        customer = LegacyCustomer(row_number=row_number)
        customer.code = code_val
        
        # Extraer campos
        def get_val(key: str) -> str:
            idx = col_map.get(key)
            if idx is not None and idx < len(row) and row[idx]:
                val = str(row[idx]).strip()
                # Limpiar .0 de floats en campos numéricos
                if key in ['zip'] and val.endswith('.0'):
                    val = val[:-2]
                return val
            return ""
        
        customer.name = get_val('name')
        customer.street = get_val('street')
        customer.city = get_val('city')
        customer.zip_code = get_val('zip')
        customer.phone = get_val('phone')
        customer.cuit = get_val('cuit')
        customer.iva_type = get_val('iva')
        customer.email = get_val('email')
        
        # Validaciones
        if not customer.name:
            customer.is_valid = False
            customer.validation_errors.append("Sin nombre")
        
        # Limpiar CUIT inválido
        if customer.cuit:
            cuit_clean = re.sub(r'[^0-9]', '', customer.cuit)
            if len(cuit_clean) < 10:
                customer.cuit = ""  # CUIT inválido, limpiar
        
        return customer


# =============================================================================
# CLIENTE ODOO (XML-RPC)
# =============================================================================

class OdooClient:
    """Cliente XML-RPC para Odoo"""
    
    def __init__(self, url: str, db: str, user: str, password: str):
        self.url = url.rstrip("/")
        self.db = db
        self.user = user
        self.password = password
        
        logger.info(f"Conectando a Odoo: {self.url}")
        
        common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common")
        self.uid = common.authenticate(self.db, self.user, self.password, {})
        
        if not self.uid:
            raise RuntimeError(f"Error de autenticación en Odoo. Usuario: {self.user}")
        
        self.models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object")
        logger.info(f"Conectado exitosamente. UID: {self.uid}")
    
    def execute_kw(self, model: str, method: str, args: List, kwargs: Optional[Dict] = None):
        kwargs = kwargs or {}
        return self.models.execute_kw(
            self.db, self.uid, self.password, model, method, args, kwargs
        )
    
    def search(self, model: str, domain: List, limit: int = 0) -> List[int]:
        opts = {"limit": limit} if limit else {}
        return self.execute_kw(model, "search", [domain], opts)
    
    def search_read(self, model: str, domain: List, fields: List[str], limit: int = 0) -> List[Dict]:
        opts = {"fields": fields}
        if limit:
            opts["limit"] = limit
        return self.execute_kw(model, "search_read", [domain], opts)
    
    def create(self, model: str, vals: Dict) -> int:
        return self.execute_kw(model, "create", [vals])
    
    def write(self, model: str, ids: List[int], vals: Dict) -> bool:
        return self.execute_kw(model, "write", [ids, vals])


# =============================================================================
# IMPORTADOR DE CLIENTES
# =============================================================================

class CustomerImporter:
    """Importador de clientes a Odoo"""
    
    def __init__(self, client: OdooClient, dry_run: bool = True, update_existing: bool = False):
        self.client = client
        self.dry_run = dry_run
        self.update_existing = update_existing
        self.result = ImportResult(dry_run=dry_run)
        
        # Lock para operaciones thread-safe
        self._lock = threading.Lock()
        
        # Cache
        self._iva_type_cache: Dict[str, int] = {}
        self._existing_by_ref: Dict[str, int] = {}
        self._existing_by_vat: Dict[str, int] = {}
        self._country_ar_id: Optional[int] = None
        self._state_misiones_id: Optional[int] = None
        self._state_caba_id: Optional[int] = None
        self._state_bsas_id: Optional[int] = None
        self._id_type_cuit: Optional[int] = None
        self._id_type_dni: Optional[int] = None
        self._city_to_state: Dict[str, int] = {}
    
    def import_customers(self, customers: List[LegacyCustomer], num_threads: int = NUM_THREADS) -> ImportResult:
        """Importa los clientes a Odoo usando múltiples hilos"""
        logger.info(f"{'[DRY-RUN] ' if self.dry_run else ''}Iniciando importación de {len(customers)} clientes con {num_threads} hilos")
        
        self.result.total_rows = len(customers)
        
        try:
            self._init_cache()
            
            # Filtrar clientes válidos
            valid_customers = []
            for customer in customers:
                if not customer.is_valid:
                    self.result.invalid_customers += 1
                else:
                    self.result.valid_customers += 1
                    valid_customers.append(customer)
            
            # Procesar en paralelo
            if self.dry_run:
                # En dry-run, procesar secuencialmente para mantener orden del log
                for customer in valid_customers:
                    self._process_customer(customer)
            else:
                # En ejecución real, usar ThreadPoolExecutor
                with ThreadPoolExecutor(max_workers=num_threads) as executor:
                    futures = {executor.submit(self._process_customer_threaded, c): c for c in valid_customers}
                    for future in as_completed(futures):
                        customer = futures[future]
                        try:
                            future.result()
                        except Exception as e:
                            logger.error(f"Error procesando {customer.name}: {e}")
            
            self._log_summary()
            
        except Exception as e:
            logger.exception("Error durante la importación")
            self.result.errors.append(f"Error crítico: {str(e)}")
        
        return self.result
    
    def _process_customer_threaded(self, customer: LegacyCustomer):
        """Wrapper thread-safe para procesar cliente"""
        # Cada hilo necesita su propia conexión XML-RPC
        thread_client = OdooClient(
            ODOO_URL, ODOO_DB, ODOO_USER, ODOO_PASSWORD
        )
        self._process_customer_with_client(customer, thread_client)
    
    def _init_cache(self):
        """Inicializa caches de datos existentes"""
        # Obtener país Argentina
        countries = self.client.search_read(
            "res.country",
            [("code", "=", "AR")],
            ["id"]
        )
        if countries:
            self._country_ar_id = countries[0]["id"]
            logger.info(f"País Argentina: ID {self._country_ar_id}")
        
        # Cargar tipos de responsabilidad AFIP
        iva_types = self.client.search_read(
            "l10n_ar.afip.responsibility.type",
            [],
            ["id", "name"]
        )
        for iva in iva_types:
            self._iva_type_cache[iva["name"]] = iva["id"]
        logger.info(f"Tipos IVA cargados: {len(self._iva_type_cache)}")
        
        # Cargar tipos de identificación (CUIT, DNI)
        id_types = self.client.search_read(
            "l10n_latam.identification.type",
            [],
            ["id", "name"]
        )
        for id_type in id_types:
            if id_type["name"] == "CUIT":
                self._id_type_cuit = id_type["id"]
            elif id_type["name"] == "DNI":
                self._id_type_dni = id_type["id"]
        logger.info(f"Tipo CUIT: ID {self._id_type_cuit}, Tipo DNI: ID {self._id_type_dni}")
        
        # Cargar provincias de Argentina
        states = self.client.search_read(
            "res.country.state",
            [("country_id", "=", self._country_ar_id)],
            ["id", "name", "code"]
        )
        for state in states:
            if state["name"] == "Misiones":
                self._state_misiones_id = state["id"]
            elif state["name"] == "Ciudad Autónoma de Buenos Aires":
                self._state_caba_id = state["id"]
            elif state["name"] == "Buenos Aires":
                self._state_bsas_id = state["id"]
        logger.info(f"Provincia Misiones: ID {self._state_misiones_id}")
        
        # Mapeo de ciudades a provincias
        self._city_to_state = {
            # Misiones
            "PUERTO IGUAZU": self._state_misiones_id,
            "PUERTO IGUAZÚ": self._state_misiones_id,
            "PTO IGUAZU": self._state_misiones_id,
            "PTO. IGUAZU": self._state_misiones_id,
            "PUERTO IGIAZU": self._state_misiones_id,
            "IGUAZU": self._state_misiones_id,
            "POSADAS": self._state_misiones_id,
            "POSADAS MISIONES": self._state_misiones_id,
            "OBERA": self._state_misiones_id,
            "OBERÁ": self._state_misiones_id,
            "ELDORADO": self._state_misiones_id,
            "WANDA": self._state_misiones_id,
            "COLONIA WANDA": self._state_misiones_id,
            "PUERTO ESPERANZA": self._state_misiones_id,
            "PUERTO LIBERTAD": self._state_misiones_id,
            "ARISTOBULO DEL VALLE": self._state_misiones_id,
            "JARDIN AMERICA": self._state_misiones_id,
            "JARDÍN AMERICA": self._state_misiones_id,
            "LEANDRO N. ALEM": self._state_misiones_id,
            "GARUPA": self._state_misiones_id,
            "GARUPÁ": self._state_misiones_id,
            "MISIONES": self._state_misiones_id,
            "MONTECARLO": self._state_misiones_id,
            "COMANDANTE ANDRESITO": self._state_misiones_id,
            "ANDRESITO": self._state_misiones_id,
            # CABA
            "CABA": self._state_caba_id,
            "CIUDAD AUTONOMA DE BUENOS AIRES": self._state_caba_id,
            "CIUDAD AUTONOMA DE BUENOS": self._state_caba_id,
            "CAPITAL FEDERAL": self._state_caba_id,
            # Buenos Aires
            "BUENOS AIRES": self._state_bsas_id,
            "VILLA BALLESTER": self._state_bsas_id,
            "MAR DEL PLATA": self._state_bsas_id,
            "MAR DEL PLATA NORTE": self._state_bsas_id,
            "MAR DEL PLATA SUR": self._state_bsas_id,
            "LOMA HERMOSA": self._state_bsas_id,
        }
        
        # Cargar partners existentes por ref
        partners = self.client.search_read(
            "res.partner",
            [("ref", "!=", False)],
            ["id", "ref", "vat"]
        )
        for p in partners:
            if p.get("ref"):
                self._existing_by_ref[str(p["ref"])] = p["id"]
            if p.get("vat"):
                vat_clean = re.sub(r'[^0-9]', '', p["vat"])
                self._existing_by_vat[vat_clean] = p["id"]
        
        logger.info(f"Partners existentes por ref: {len(self._existing_by_ref)}")
        logger.info(f"Partners existentes por VAT: {len(self._existing_by_vat)}")
    
    def _get_iva_type_id(self, iva_code: str) -> Optional[int]:
        """Obtiene ID del tipo de responsabilidad AFIP"""
        if not iva_code:
            return None
        
        iva_code_upper = iva_code.upper().strip()
        iva_name = IVA_TYPE_MAP.get(iva_code_upper)
        
        if not iva_name:
            return None
        
        return self._iva_type_cache.get(iva_name)
    
    def _get_state_from_city(self, city: str) -> Optional[int]:
        """Obtiene ID de provincia basándose en la ciudad"""
        if not city:
            return self._state_misiones_id  # Default: Misiones
        
        city_upper = city.upper().strip()
        
        # Buscar en mapeo directo
        if city_upper in self._city_to_state:
            return self._city_to_state[city_upper]
        
        # Buscar parcialmente
        for mapped_city, state_id in self._city_to_state.items():
            if mapped_city in city_upper or city_upper in mapped_city:
                return state_id
        
        # Default: Misiones (la mayoría de clientes son de ahí)
        return self._state_misiones_id
    
    def _find_existing_partner(self, customer: LegacyCustomer) -> Optional[int]:
        """Busca si el partner ya existe"""
        # Buscar por ref (código legacy)
        if customer.code in self._existing_by_ref:
            return self._existing_by_ref[customer.code]
        
        # Buscar por CUIT
        if customer.cuit:
            cuit_clean = re.sub(r'[^0-9]', '', customer.cuit)
            if cuit_clean in self._existing_by_vat:
                return self._existing_by_vat[cuit_clean]
        
        return None
    
    def _process_customer(self, customer: LegacyCustomer):
        """Procesa un cliente individual (usa cliente por defecto)"""
        self._process_customer_with_client(customer, self.client)
    
    def _process_customer_with_client(self, customer: LegacyCustomer, client: 'OdooClient'):
        """Procesa un cliente individual con un cliente XML-RPC específico"""
        # Check de existencia es thread-safe gracias al lock
        with self._lock:
            existing_id = self._find_existing_partner(customer)
        
        if existing_id:
            if self.update_existing:
                self._update_customer_with_client(customer, existing_id, client)
            else:
                self._log(f"SKIP (existe): {customer.code} - {customer.name} (ID: {existing_id})")
                with self._lock:
                    self.result.customers_skipped += 1
            return
        
        # Crear nuevo
        self._create_customer_with_client(customer, client)
    
    def _create_customer(self, customer: LegacyCustomer):
        """Crea un nuevo cliente en Odoo (usa cliente por defecto)"""
        self._create_customer_with_client(customer, self.client)
    
    def _create_customer_with_client(self, customer: LegacyCustomer, client: 'OdooClient'):
        """Crea un nuevo cliente en Odoo con un cliente XML-RPC específico"""
        vals = {
            "name": customer.name,
            "ref": customer.code,
            "customer_rank": 1,
            "is_company": self._is_company(customer.name),
            "company_id": False,  # Partner compartido
        }
        
        if customer.street:
            vals["street"] = customer.street
        
        if customer.city:
            vals["city"] = customer.city
        
        if customer.zip_code:
            vals["zip"] = customer.zip_code
        
        if customer.phone:
            vals["phone"] = customer.phone
        
        # VAT sin guiones y tipo de identificación
        if customer.vat_clean:
            vals["vat"] = customer.vat_clean
            # Asignar tipo de identificación: CUIT o DNI
            if customer.is_cuit and self._id_type_cuit:
                vals["l10n_latam_identification_type_id"] = self._id_type_cuit
            elif customer.is_dni and self._id_type_dni:
                vals["l10n_latam_identification_type_id"] = self._id_type_dni
        
        if customer.email and '@' in customer.email:
            vals["email"] = customer.email
        
        if self._country_ar_id:
            vals["country_id"] = self._country_ar_id
        
        # Provincia basada en ciudad
        state_id = self._get_state_from_city(customer.city)
        if state_id:
            vals["state_id"] = state_id
        
        # Tipo de responsabilidad AFIP
        iva_type_id = self._get_iva_type_id(customer.iva_type)
        if iva_type_id:
            vals["l10n_ar_afip_responsibility_type_id"] = iva_type_id
        
        if self.dry_run:
            self._log(f"[DRY-RUN] Crearía: {customer.code} - {customer.name}")
            with self._lock:
                self.result.customers_created += 1
        else:
            try:
                partner_id = client.create("res.partner", vals)
                with self._lock:
                    self.result.created_ids.append(partner_id)
                    self.result.customers_created += 1
                    # Actualizar cache
                    self._existing_by_ref[customer.code] = partner_id
                    if customer.cuit:
                        cuit_clean = re.sub(r'[^0-9]', '', customer.cuit)
                        self._existing_by_vat[cuit_clean] = partner_id
                self._log(f"Creado: {customer.code} - {customer.name} (ID: {partner_id})")
                    
            except Exception as e:
                with self._lock:
                    self.result.errors.append(f"Error creando {customer.name}: {str(e)}")
                logger.error(f"Error creando {customer.name}: {e}")
    
    def _update_customer(self, customer: LegacyCustomer, partner_id: int):
        """Actualiza un cliente existente (usa cliente por defecto)"""
        self._update_customer_with_client(customer, partner_id, self.client)
    
    def _update_customer_with_client(self, customer: LegacyCustomer, partner_id: int, client: 'OdooClient'):
        """Actualiza un cliente existente con un cliente XML-RPC específico"""
        vals = {}
        
        if customer.street:
            vals["street"] = customer.street
        if customer.city:
            vals["city"] = customer.city
        if customer.zip_code:
            vals["zip"] = customer.zip_code
        if customer.phone:
            vals["phone"] = customer.phone
        if customer.email and '@' in customer.email:
            vals["email"] = customer.email
        
        iva_type_id = self._get_iva_type_id(customer.iva_type)
        if iva_type_id:
            vals["l10n_ar_afip_responsibility_type_id"] = iva_type_id
        
        if not vals:
            self._log(f"SKIP (sin cambios): {customer.code} - {customer.name}")
            with self._lock:
                self.result.customers_skipped += 1
            return
        
        if self.dry_run:
            self._log(f"[DRY-RUN] Actualizaría: {customer.code} - {customer.name}")
            with self._lock:
                self.result.customers_updated += 1
        else:
            try:
                client.write("res.partner", [partner_id], vals)
                with self._lock:
                    self.result.customers_updated += 1
                self._log(f"Actualizado: {customer.code} - {customer.name} (ID: {partner_id})")
            except Exception as e:
                with self._lock:
                    self.result.errors.append(f"Error actualizando {customer.name}: {str(e)}")
    
    def _is_company(self, name: str) -> bool:
        """Determina si es empresa basándose en el nombre"""
        if not name:
            return False
        name_upper = name.upper()
        company_indicators = ['S.A.', 'SA ', 'SRL', 'S.R.L.', 'SAS', 'S.A.S.', 
                             'SOCIEDAD', 'EMPRESA', 'CIA', 'COMPAÑIA', 'LTDA',
                             'S.C.', 'S.H.', 'COOPERATIVA', 'FUNDACION']
        return any(ind in name_upper for ind in company_indicators)
    
    def _log(self, message: str):
        """Registra mensaje"""
        logger.info(message)
        self.result.log_entries.append(message)
    
    def _log_summary(self):
        """Registra resumen"""
        summary = [
            "",
            "=" * 60,
            f"{'[DRY-RUN] ' if self.dry_run else ''}RESUMEN DE IMPORTACIÓN",
            "=" * 60,
            f"Total filas procesadas:   {self.result.total_rows}",
            f"Clientes válidos:         {self.result.valid_customers}",
            f"Clientes inválidos:       {self.result.invalid_customers}",
            f"Clientes creados:         {self.result.customers_created}",
            f"Clientes actualizados:    {self.result.customers_updated}",
            f"Clientes omitidos:        {self.result.customers_skipped}",
            f"Errores:                  {len(self.result.errors)}",
            "=" * 60,
        ]
        
        for line in summary:
            logger.info(line)
            self.result.log_entries.append(line)


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Importación de clientes legacy a Odoo",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  # Solo analizar el Excel
  python import_legacy_customers.py --excel /path/to/clientes.xlsx --parse-only
  
  # Dry-run (simular)
  python import_legacy_customers.py --excel /path/to/clientes.xlsx --dry-run
  
  # Ejecutar importación real
  python import_legacy_customers.py --excel /path/to/clientes.xlsx --execute
  
  # Actualizar existentes
  python import_legacy_customers.py --excel /path/to/clientes.xlsx --execute --update-existing
        """
    )
    
    parser.add_argument("--excel", "-e", required=True, help="Ruta al Excel de clientes")
    parser.add_argument("--parse-only", "-p", action="store_true", help="Solo parsear, no conectar a Odoo")
    parser.add_argument("--dry-run", "-d", action="store_true", help="Simular importación")
    parser.add_argument("--execute", "-x", action="store_true", help="Ejecutar importación real")
    parser.add_argument("--update-existing", "-u", action="store_true", help="Actualizar clientes existentes")
    parser.add_argument("--url", default=ODOO_URL, help="URL de Odoo")
    parser.add_argument("--db", default=ODOO_DB, help="Base de datos Odoo")
    parser.add_argument("--user", default=ODOO_USER, help="Usuario Odoo")
    parser.add_argument("--password", default=ODOO_PASSWORD, help="Contraseña Odoo")
    parser.add_argument("--verbose", "-v", action="store_true", help="Modo verbose")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    if not args.parse_only and not args.dry_run and not args.execute:
        parser.error("Debe especificar --parse-only, --dry-run o --execute")
    
    if not os.path.exists(args.excel):
        parser.error(f"El archivo no existe: {args.excel}")
    
    # 1. Parsear Excel
    print("\n" + "=" * 60)
    print("PASO 1: ANÁLISIS DEL ARCHIVO EXCEL")
    print("=" * 60 + "\n")
    
    parser_obj = LegacyCustomerParser(args.excel)
    customers = parser_obj.parse()
    
    valid = [c for c in customers if c.is_valid]
    invalid = [c for c in customers if not c.is_valid]
    
    print(f"\n📊 Archivo:           {args.excel}")
    print(f"📄 Total filas:       {len(customers)}")
    print(f"✅ Clientes válidos:  {len(valid)}")
    print(f"❌ Clientes inválidos:{len(invalid)}")
    
    # Estadísticas de campos
    with_cuit = len([c for c in valid if c.cuit])
    with_email = len([c for c in valid if c.email and '@' in c.email])
    with_phone = len([c for c in valid if c.phone])
    
    print(f"\n📈 ESTADÍSTICAS:")
    print(f"   Con CUIT:     {with_cuit} ({100*with_cuit/len(valid):.1f}%)")
    print(f"   Con Email:    {with_email} ({100*with_email/len(valid):.1f}%)")
    print(f"   Con Teléfono: {with_phone} ({100*with_phone/len(valid):.1f}%)")
    
    # Muestra
    print("\n📋 MUESTRA DE CLIENTES:")
    print("-" * 80)
    for c in valid[:5]:
        print(f"  {c.code:>6} | {c.name[:35]:<35} | {c.vat_clean or 'Sin CUIT':<15} | {c.iva_type}")
    if len(valid) > 5:
        print(f"  ... y {len(valid) - 5} más")
    
    if args.parse_only:
        print("\n✅ Análisis completado (modo --parse-only)")
        sys.exit(0)
    
    # 2. Importar a Odoo
    print("\n" + "=" * 60)
    print("PASO 2: IMPORTACIÓN A ODOO")
    print("=" * 60 + "\n")
    
    try:
        client = OdooClient(args.url, args.db, args.user, args.password)
    except Exception as e:
        print(f"❌ Error conectando a Odoo: {e}")
        sys.exit(1)
    
    importer = CustomerImporter(
        client=client,
        dry_run=args.dry_run,
        update_existing=args.update_existing
    )
    
    result = importer.import_customers(valid)
    
    # Resultados
    print("\n" + "=" * 60)
    print(f"{'[DRY-RUN] ' if args.dry_run else ''}RESULTADO DE LA IMPORTACIÓN")
    print("=" * 60)
    print(f"✅ Clientes creados:      {result.customers_created}")
    print(f"🔄 Clientes actualizados: {result.customers_updated}")
    print(f"⏭️  Clientes omitidos:     {result.customers_skipped}")
    
    if result.errors:
        print(f"\n❌ ERRORES ({len(result.errors)}):")
        for e in result.errors[:10]:
            print(f"   - {e}")
    
    if args.dry_run:
        print("\n💡 Este fue un DRY-RUN. Para ejecutar realmente, use --execute")
    else:
        print("\n✅ Importación completada exitosamente")
        if result.created_ids:
            print(f"   IDs creados: {result.created_ids[:20]}")
            if len(result.created_ids) > 20:
                print(f"   ... y {len(result.created_ids) - 20} más")


if __name__ == "__main__":
    main()
